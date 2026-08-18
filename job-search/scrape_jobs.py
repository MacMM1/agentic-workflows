"""Scrape a Find a Job (jobs.service.gov.uk) search into an Excel sheet.

Usage:
    python job-search/scrape_jobs.py --url "<search results URL>" [--output PATH]
        [--per-page 30] [--concurrency 3] [--rate-limit 12] [--limit N] [--yes]
"""

import argparse
import os
import re
import sys
import threading
import time
from collections import deque
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlsplit, urlunsplit, parse_qs, urlencode

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

API_BASE = "https://api.firecrawl.dev/v2/scrape"
JOB_URL_RE = re.compile(r"^https://www\.jobs\.service\.gov\.uk/jobs/[0-9a-f]{24}$")
TOTAL_RESULTS_RE = re.compile(r"of\s+\*{0,2}([\d,]+)\*{0,2}")
RETRY_AFTER_RE = re.compile(r"retry after (\d+)s")


class RateLimiter:
    """Sliding-window limiter so concurrent workers stay under the account's req/min cap."""

    def __init__(self, max_per_minute):
        self.max_per_minute = max_per_minute
        self.lock = threading.Lock()
        self.timestamps = deque()

    def acquire(self):
        while True:
            with self.lock:
                now = time.monotonic()
                while self.timestamps and now - self.timestamps[0] > 60:
                    self.timestamps.popleft()
                if len(self.timestamps) < self.max_per_minute:
                    self.timestamps.append(now)
                    return
                wait = 60 - (now - self.timestamps[0]) + 0.1
            time.sleep(wait)

COLUMNS = [
    ("title", "Job Title"),
    ("employer", "Employer"),
    ("location", "Location"),
    ("salary", "Salary"),
    ("contract_type", "Contract Type"),
    ("hours", "Hours"),
    ("working_pattern", "Working Pattern"),
    ("posting_date", "Posting Date"),
    ("closing_date", "Closing Date"),
    ("simplified_description", "Simplified Description"),
    ("job_url", "Job URL"),
    ("apply_url", "Apply URL"),
]

JOB_SCHEMA = {
    "type": "object",
    "properties": {
        "title": {"type": "string"},
        "employer": {"type": "string"},
        "location": {"type": "string"},
        "salary": {"type": "string"},
        "contract_type": {"type": "string"},
        "hours": {"type": "string"},
        "working_pattern": {"type": "string"},
        "posting_date": {"type": "string"},
        "closing_date": {"type": "string"},
        "simplified_description": {"type": "string"},
    },
    "required": ["title", "employer", "simplified_description"],
}

JOB_PROMPT = (
    "Extract the job posting fields from the table at the top of the page. "
    "For simplified_description, write a 2-3 sentence plain-language summary of the "
    "role's main duties and requirements, excluding company background, benefits "
    "lists, and safeguarding/legal boilerplate. Leave a field as an empty string if "
    "it is not present on the page."
)


def api_key():
    load_dotenv()
    key = os.environ.get("FIRECRAWL_API_KEY")
    if not key:
        sys.exit("FIRECRAWL_API_KEY is not set in .env — see job-search/README.md setup.")
    return key


def request_with_retries(session, body, headers, rate_limiter, retries=5):
    backoff = 2
    for attempt in range(1, retries + 1):
        rate_limiter.acquire()
        try:
            resp = session.post(API_BASE, json=body, headers=headers, timeout=60)
        except requests.RequestException as exc:
            if attempt == retries:
                raise
            time.sleep(backoff)
            backoff *= 2
            continue
        if resp.status_code == 429:
            if attempt == retries:
                raise RuntimeError(f"429 after {retries} attempts: {resp.text[:300]}")
            match = RETRY_AFTER_RE.search(resp.text)
            time.sleep(int(match.group(1)) + 1 if match else backoff)
            backoff *= 2
            continue
        if resp.status_code >= 500:
            if attempt == retries:
                raise RuntimeError(f"{resp.status_code} after {retries} attempts: {resp.text[:300]}")
            time.sleep(backoff)
            backoff *= 2
            continue
        if not resp.ok:
            raise RuntimeError(f"{resp.status_code} error: {resp.text[:300]}")
        return resp.json()
    raise RuntimeError("unreachable")


def build_listing_url(base_url, page_number, per_page):
    parts = urlsplit(base_url)
    qs = parse_qs(parts.query, keep_blank_values=True)
    qs["pageNumber"] = [str(page_number)]
    qs["resultsPerPage"] = [str(per_page)]
    query = urlencode(qs, doseq=True)
    return urlunsplit((parts.scheme, parts.netloc, parts.path, query, ""))


def extract_job_links(links):
    seen = dict.fromkeys(u for u in links if JOB_URL_RE.match(u))
    return list(seen)


def discover_job_urls(session, headers, rate_limiter, search_url, per_page):
    first_page_url = build_listing_url(search_url, 1, per_page)
    data = request_with_retries(
        session,
        {"url": first_page_url, "formats": ["markdown", "links"], "onlyMainContent": True},
        headers,
        rate_limiter,
    )["data"]

    match = TOTAL_RESULTS_RE.search(data.get("markdown", ""))
    if not match:
        sys.exit("Could not find a result count on the search page — check the URL is a valid Find a Job search.")
    total_results = int(match.group(1).replace(",", ""))

    job_urls = extract_job_links(data.get("links", []))
    total_pages = -(-total_results // per_page)  # ceil div

    print(f"Found {total_results} results across {total_pages} page(s).")

    for page in range(2, total_pages + 1):
        page_url = build_listing_url(search_url, page, per_page)
        page_data = request_with_retries(
            session, {"url": page_url, "formats": ["links"], "onlyMainContent": True}, headers, rate_limiter
        )["data"]
        job_urls.extend(extract_job_links(page_data.get("links", [])))
        print(f"  scanned page {page}/{total_pages} — {len(job_urls)} job links so far", flush=True)

    deduped = list(dict.fromkeys(job_urls))
    return deduped, total_results


def scrape_job(session, headers, rate_limiter, job_url):
    body = {
        "url": job_url,
        "onlyMainContent": True,
        "formats": [{"type": "json", "prompt": JOB_PROMPT, "schema": JOB_SCHEMA}],
    }
    payload = request_with_retries(session, body, headers, rate_limiter)
    data = payload.get("data", payload)
    extracted = data.get("json") or {}
    row = {key: extracted.get(key, "") or "" for key, _ in COLUMNS}
    row["job_url"] = job_url
    row["apply_url"] = job_url + "/apply"
    return row


def write_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"
    ws.append([label for _, label in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    hyperlink_font = Font(color="0563C1", underline="single")
    for row in rows:
        ws.append([row[key] for key, _ in COLUMNS])
        excel_row = ws.max_row
        for col_index, (key, _) in enumerate(COLUMNS, start=1):
            if key in ("job_url", "apply_url") and row[key]:
                cell = ws.cell(row=excel_row, column=col_index)
                cell.hyperlink = row[key]
                cell.font = hyperlink_font

    widths = {
        "Job Title": 32, "Employer": 26, "Location": 24, "Salary": 20,
        "Contract Type": 14, "Hours": 12, "Working Pattern": 16,
        "Posting Date": 13, "Closing Date": 13, "Simplified Description": 60,
        "Job URL": 45, "Apply URL": 45,
    }
    for col_index, (_, label) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = widths.get(label, 20)

    ws.auto_filter.ref = ws.dimensions
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)


def slugify(url):
    qs = parse_qs(urlsplit(url).query)
    keywords = qs.get("keywords", [""])[0]
    location = qs.get("location", [""])[0]
    parts = [p for p in (keywords, location) if p]
    slug = "-".join(parts) or "jobs"
    return re.sub(r"[^a-zA-Z0-9-]+", "-", slug).strip("-").lower()


def main():
    parser = argparse.ArgumentParser(description="Scrape a Find a Job search into an Excel sheet.")
    parser.add_argument("--url", required=True, help="Find a Job search results URL")
    parser.add_argument("--output", help="Output .xlsx path")
    parser.add_argument("--per-page", type=int, default=30, help="Results per listing page (max 30)")
    parser.add_argument("--concurrency", type=int, default=3, help="Parallel job-page scrapes")
    parser.add_argument("--rate-limit", type=int, default=12, help="Max Firecrawl requests per minute (stay under your plan's cap)")
    parser.add_argument("--limit", type=int, help="Cap the number of jobs scraped (for testing)")
    parser.add_argument("--yes", action="store_true", help="Skip the cost confirmation prompt")
    args = parser.parse_args()

    key = api_key()
    headers = {"Authorization": f"Bearer {key}", "Content-Type": "application/json"}
    session = requests.Session()
    rate_limiter = RateLimiter(args.rate_limit)

    job_urls, total_results = discover_job_urls(session, headers, rate_limiter, args.url, args.per_page)
    if args.limit:
        job_urls = job_urls[: args.limit]

    listing_credits = -(-total_results // args.per_page)
    detail_credits = len(job_urls) * 5
    print(
        f"\n{len(job_urls)} job pages to scrape. Estimated Firecrawl credits: "
        f"~{listing_credits + detail_credits} ({listing_credits} listing + {detail_credits} detail)."
    )
    if not args.yes:
        answer = input("Continue? [y/N] ").strip().lower()
        if answer != "y":
            print("Aborted.")
            return

    rows = []
    failures = []
    with ThreadPoolExecutor(max_workers=args.concurrency) as pool:
        futures = {pool.submit(scrape_job, session, headers, rate_limiter, url): url for url in job_urls}
        done = 0
        for future in as_completed(futures):
            url = futures[future]
            done += 1
            try:
                row = future.result()
                rows.append(row)
                print(f"  [{done}/{len(job_urls)}] {row['title'] or url}", flush=True)
            except Exception as exc:
                failures.append((url, str(exc)))
                print(f"  [{done}/{len(job_urls)}] FAILED {url}: {exc}", flush=True)

    output_path = args.output or os.path.join("job-search-results", f"{slugify(args.url)}.xlsx")
    write_excel(rows, output_path)

    print(f"\nWrote {len(rows)} rows to {output_path}")
    if failures:
        print(f"{len(failures)} job(s) failed and were skipped:")
        for url, error in failures:
            print(f"  - {url}: {error}")


if __name__ == "__main__":
    main()
